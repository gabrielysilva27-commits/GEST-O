from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ALLOWED_RESPONSIBLES = {
    "Diego": "Diego Teixeira",
    "Nathan": "Nathan Capitão",
    "Iago": "Iago Rodrigues",
}


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Uso: import-audit-actions.py origem.xlsx destino.js")

    source = Path(sys.argv[1])
    destination = Path(sys.argv[2])
    sheet = load_workbook(source, data_only=False, read_only=True)["Checklist"]
    actions = []

    for excel_row, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        pilar, identifier, block, question, action, target, responsible, *_rest = row
        if identifier is None or str(responsible).strip() not in ALLOWED_RESPONSIBLES:
            continue

        owner_key = str(responsible).strip()
        ok_mark = row[9]
        actions.append(
            {
                "id": int(identifier),
                "sourceRow": excel_row,
                "pilar": str(pilar).strip(),
                "bloco": str(block).strip(),
                "questao": str(question).strip(),
                "acao": str(action).strip(),
                "meta": str(target).strip(),
                "responsavel": ALLOWED_RESPONSIBLES[owner_key],
                "username": owner_key,
                "status": "done" if ok_mark == "☑" else "pending",
            }
        )

    payload = json.dumps(actions, ensure_ascii=False, indent=2)
    destination.write_text(
        "// Gerado de ACOMPANHAMENTO AÇÕES - AUDITORIA.xlsx.\n"
        f"export const AUDIT_ACTIONS = {payload};\n",
        encoding="utf-8",
    )
    print(f"{len(actions)} ações importadas para {destination}")


if __name__ == "__main__":
    main()
